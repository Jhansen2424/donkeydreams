# Deep audit v2 — closes the blind spots of audit-import.py:
#   1. (title, date) PAIR-level reconciliation of every deworming/vaccination
#      entry against the words adjacent to each date in the cell (multiset —
#      catches a missing second entry on the same date, wrong drug names, and
#      invented entries).
#   2. Relationship links (parents / children / bondedWith): every linked name
#      must literally appear in that donkey's Relationships cell.
#   3. Trim visit DATES match the date written in the Trim History cell.
#   4. The ignored "Vaccinated" column: every date in it must already be
#      covered by an imported entry (proves ignoring it lost nothing).
#   5. Prints every relationship extraction for manual review.
#
# Usage: python scripts/audit-import-v2.py "<adoption xlsx>" "<checklist xlsx>"
import json
import re
import sys
from collections import Counter

import openpyxl

ADOPTION_XLSX, CHECKLIST_XLSX = sys.argv[1], sys.argv[2]
issues = []

OVERRIDES = {
    "JACK JACK": "Jack Jack", "DANNY BOY": "Danny Boy",
    "ISABELLA (IZZY)": "Izabella (Izzy)", "SKYLA (SKYE)": "Skyla (Skye)",
    "NELLY BELLE": "Nelly Belle", "ELENORA": "Elanora",
}

def resolve(name):
    u = name.strip().upper()
    if u in OVERRIDES:
        return OVERRIDES[u]
    return re.sub(r"\b\w", lambda m: m.group(0).upper(), name.strip().lower())

def iso(val):
    if val is None:
        return None
    if hasattr(val, "year"):
        return f"{val.year:04d}-{val.month:02d}-{val.day:02d}"
    m = re.match(r"^\s*(?:born\s+)?(\d{1,2})/(\d{1,2})/(\d{2,4})\s*$", str(val), re.I)
    if not m:
        return None
    y = m.group(3)
    if len(y) == 2:
        y = "20" + y
    return f"{y}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"

def txt(val):
    return re.sub(r"\s+", " ", str(val)).strip() if val is not None else ""

# ── Load generated data ──
profiles_src = open("src/lib/donkey-profiles-data.ts", encoding="utf-8").read()
profiles = {}
for m in re.finditer(r'^  \["(?:[^"\\]|\\.)*", (\{.*\})\],$', profiles_src, re.M):
    p = json.loads(m.group(1))
    profiles[p["name"]] = p

def ts_entries(src, array_name):
    block = re.search(rf"export const {array_name}[^=]*= \[\n(.*?)\n\];", src, re.S)
    out = []
    if not block:
        return out
    for line in block.group(1).split("\n"):
        fields = {}
        for key in ("id", "animal", "type", "title", "date", "description", "notes"):
            fm = re.search(rf'{key}: ("(?:[^"\\]|\\.)*")', line)
            if fm:
                fields[key] = json.loads(fm.group(1))
        if fields:
            out.append(fields)
    return out

trim_visits = ts_entries(profiles_src, "adoptionTrimVisits")
dw_src = open("src/lib/deworming-vaccination-data.ts", encoding="utf-8").read()
dw_entries = ts_entries(dw_src, "importedDewormingEntries")
vx_entries = ts_entries(dw_src, "importedVaccinationEntries")

# ── Title token normalization (independent of the parser's cleanTitle) ──
STOP = {"of", "and", "on", "the", "lot"}

def norm_tokens(s):
    s = s.lower()
    s = s.replace("rabiies", "rabies")
    s = re.sub(r"\bvermectin\b", "ivermectin", s)
    s = re.sub(r"\bdewormed\b", "deworming", s)
    s = re.sub(r"\(([a-z0-9-]{4,})\)", " ", s)  # lot numbers
    s = re.sub(r"[^a-z0-9]+", " ", s)  # '+' is a separator ("+VEE" == "+ VEE")
    toks = [t for t in s.split() if t not in STOP]
    return frozenset(toks) if toks else frozenset()

DATE_RE = re.compile(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?")

def expected_pairs(cell, fallback, default_title):
    """Independent re-derivation of (title-tokens, date) pairs from a cell.
    Semicolons are hard breaks (bare date after one → default title);
    commas are soft (a title carries across comma-separated bare dates)."""
    pairs = []
    if not cell:
        return pairs
    text = re.sub(r"\s+", " ", cell).strip()
    for major in text.split(";"):
      prev_raw = None
      for chunk in major.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        toks = list(DATE_RE.finditer(chunk))
        if not toks:
            continue  # dateless fragments (lots) checked implicitly via titles
        segs = [chunk[: toks[0].start()]]
        for i, t in enumerate(toks):
            nxt = toks[i + 1].start() if i + 1 < len(toks) else len(chunk)
            segs.append(chunk[t.end():nxt])
        date_first = segs[0].strip() == ""
        for i, t in enumerate(toks):
            y = t.group(3)
            if not y:
                y = next((n.group(3) for n in toks[i + 1:] if n.group(3)), None) or (fallback[:4] if fallback else None)
                if not y:
                    continue
            if len(y) == 2:
                y = "20" + y
            date = f"{y}-{int(t.group(1)):02d}-{int(t.group(2)):02d}"
            raw = segs[i + 1] if date_first else segs[i]
            # lowercase leading context belongs to the previous event
            raw = re.sub(r"^[\s.]*(?:for|due to|because of)\s+[a-z][^A-Z]*", " ", raw)
            if not re.sub(r"\b(?:and|on|the|of)\b", " ", raw).strip(" .,"):
                raw = prev_raw if prev_raw else default_title
            else:
                prev_raw = raw
            # mirror the importer: "X and Dewormed" / "A and B" split into
            # separate events on the same date
            parts = re.split(r"\s+and\s+", raw, flags=re.I)
            for part in parts:
                title = norm_tokens(part)
                if not title:
                    continue
                if title == {"dewormed"} or title == {"deworming"}:
                    title = frozenset({"deworming"})
                pairs.append((title, date))
    # collapse exact repeats the same way the importer does
    seen = set()
    out = []
    for p in pairs:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out

def entry_pairs(entries, animal):
    out = []
    for e in entries:
        if e["animal"] != animal:
            continue
        title = norm_tokens(e["title"])
        # split "x and y" style titles won't appear (importer already split)
        out.append((title, e["date"]))
    return out

def match_multisets(name, kind, want, got, cell):
    cw, cg = Counter(want), Counter(got)
    missing = list((cw - cg).elements())
    extra = list((cg - cw).elements())
    # tolerate token-subset matches (e.g. importer title "6 Way" vs cell "gold 6 way")
    for mpair in missing[:]:
        for epair in extra[:]:
            if mpair[1] == epair[1] and (mpair[0] <= epair[0] or epair[0] <= mpair[0]):
                missing.remove(mpair)
                extra.remove(epair)
                break
    for mpair in missing:
        issues.append(f"[{kind}] {name}: cell has {sorted(mpair[0])} {mpair[1]} — no matching entry | cell: {cell[:90]!r}")
    for epair in extra:
        issues.append(f"[{kind}] {name}: entry {sorted(epair[0])} {epair[1]} not derivable from cell | cell: {cell[:90]!r}")

# ════════ Checklist: pair-level reconciliation ════════
wb2 = openpyxl.load_workbook(CHECKLIST_XLSX, data_only=True)
ws2 = wb2["Donkey Dreams - Current"]
HERDS = {"Angels", "Brave", "Dragons", "Elsie", "Legacy", "Pegasus", "Pink", "Senior", "Seniors", "Unicorns"}
checked = 0
vaccinated_checked = 0
for row in ws2.iter_rows(min_row=2):
    v = [c.value for c in row]
    if v[0] is None or (txt(v[1]) not in HERDS):
        continue
    name = resolve(str(v[0]))
    checked += 1
    dw_cell, vx_cell = txt(v[3]), txt(v[5])
    dew_fb, vacc_fb = iso(v[2]), iso(v[6])

    want_dw = expected_pairs(dw_cell, dew_fb, "Deworming")
    want_vx_all = expected_pairs(vx_cell, vacc_fb, "Vaccination")
    # A "Vaccinated"-column date with no history event is imported as a
    # generic Vaccination entry — expect it.
    vacc_col_date = iso(v[4])
    if vacc_col_date and not any(d == vacc_col_date for _, d in want_vx_all):
        want_vx_all.append((norm_tokens("Vaccination"), vacc_col_date))
    # deworming peeled out of the vaccination cell moves to the DW side
    want_dw += [p for p in want_vx_all if p[0] == frozenset({"deworming"})]
    want_vx = [p for p in want_vx_all if p[0] != frozenset({"deworming"})]
    # vaccine pairs joined by 'and' in one title: importer splits — mirror it
    split_vx = []
    for title, date in want_vx:
        split_vx.append((title, date))
    match_multisets(name, "DW", want_dw, entry_pairs(dw_entries, name), dw_cell)
    match_multisets(name, "VX", split_vx, entry_pairs(vx_entries, name), vx_cell)

    # the ignored "Vaccinated" column must be covered by existing entries
    vacc_col = iso(v[4])
    if vacc_col:
        vaccinated_checked += 1
        dates = {e["date"] for e in vx_entries if e["animal"] == name}
        if vacc_col not in dates:
            issues.append(f"[VX] {name}: 'Vaccinated' column date {vacc_col} has no entry (column was ignored)")

# ════════ Adoption: relationships + trim dates ════════
wb = openpyxl.load_workbook(ADOPTION_XLSX, data_only=True)
ws = wb["Donkey Dreams - Current"]
rel_report = []
for row in ws.iter_rows(min_row=2):
    v = [c.value for c in row]
    if v[0] is None:
        continue
    name = resolve(str(v[0]))
    p = profiles.get(name)
    if not p:
        continue
    rel_cell = (txt(v[15]) + " " + txt(v[16])).strip()
    rel_norm = rel_cell.lower().replace("elenora", "elanora")
    rel_norm = re.sub(r"\bprincesss\b", "princess", rel_norm)
    rel_norm = re.sub(r"\bgrace\b", "gracie", rel_norm)  # sheet typo for Gracie
    rel_norm = re.sub(r"\bisabella\b", "izabella", rel_norm)  # canonical spelling
    rel_norm = re.sub(r"\bmaku ahini hau\b", "makuahine hau", rel_norm)
    for field in ("parents", "children", "bondedWith"):
        for linked in p[field]:
            first = linked.split(" (")[0].lower()
            if first not in rel_norm:
                issues.append(f"[REL] {name}.{field} links {linked!r} but Relationships cell is {rel_cell[:70]!r}")
    if p["parents"] or p["children"] or p["bondedWith"]:
        rel_report.append(f"  {name}: cell={rel_cell!r} -> parents={p['parents']} children={p['children']} bonded={p['bondedWith']}")

    trim_cell = txt(v[20])
    if trim_cell and trim_cell.upper() != "N/A":
        visits = [t for t in trim_visits if t["animal"] == name]
        dm = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", trim_cell)
        mn = re.search(r"(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{4})", trim_cell.lower())
        if dm:
            y = dm.group(3)
            y = ("20" + y) if len(y) == 2 else y
            want = f"{y}-{int(dm.group(1)):02d}-{int(dm.group(2)):02d}"
        elif mn:
            months = ["january","february","march","april","may","june","july","august","september","october","november","december"]
            want = f"{mn.group(2)}-{months.index(mn.group(1))+1:02d}-01"
        else:
            want = None
        if want and not any(t["date"] == want for t in visits):
            issues.append(f"[TRIM] {name}: cell date {want} but visits {[t['date'] for t in visits]}")
        if visits and txt(visits[0].get('notes','')) != trim_cell:
            issues.append(f"[TRIM] {name}: notes differ from cell: {visits[0].get('notes','')[:60]!r}")

print(f"pair-level check: {checked} checklist donkeys, {len(dw_entries)}+{len(vx_entries)} entries reconciled; 'Vaccinated' column dates checked: {vaccinated_checked}")
print(f"relationship links verified for {len(rel_report)} donkeys; trim visits: {len(trim_visits)}")
print()
if issues:
    print(f"DISCREPANCIES: {len(issues)}")
    for i in issues:
        print("  " + i)
else:
    print("DISCREPANCIES: 0 ✔")
print()
print("── Relationship extractions (manual review) ──")
print("\n".join(rel_report))
