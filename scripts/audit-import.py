# Independent audit: re-reads both FINAL xlsx sheets and cross-checks every
# cell against the generated app data (donkey-profiles-data.ts,
# deworming-vaccination-data.ts). Prints a discrepancy report.
#
# Usage: python scripts/audit-import.py "<adoption xlsx>" "<checklist xlsx>"
import json
import re
import sys

import openpyxl

ADOPTION_XLSX, CHECKLIST_XLSX = sys.argv[1], sys.argv[2]

issues = []
ok = []


def issue(msg):
    issues.append(msg)


# ── Name resolution (mirrors the app's canonical names) ──
OVERRIDES = {
    "JACK JACK": "Jack Jack",
    "DANNY BOY": "Danny Boy",
    "ISABELLA (IZZY)": "Izabella (Izzy)",
    "SKYLA (SKYE)": "Skyla (Skye)",
    "NELLY BELLE": "Nelly Belle",
    "ELENORA": "Elanora",
}


def resolve(name):
    u = name.strip().upper()
    if u in OVERRIDES:
        return OVERRIDES[u]
    return re.sub(r"\b\w", lambda m: m.group(0).upper(), name.strip().lower())


def iso(val):
    if val is None:
        return None
    if hasattr(val, "year"):
        return f"{val.year:04d}-{val.month:02d}-{val.day:02d}"
    m = re.match(r"^\s*(?:born\s+)?(\d{1,2})/(\d{1,2})/(\d{2,4})\s*$", str(val), re.I)
    if not m:
        return None
    y = m.group(3)
    if len(y) == 2:
        y = "20" + y
    return f"{y}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"


def txt(val):
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip()


HERD_MAP = {"Elsie": "Elsie's Herd", "Pink": "Pinky's Herd", "Senior": "Seniors"}

# ── Load generated donkey-profiles-data.ts ──
profiles_src = open("src/lib/donkey-profiles-data.ts", encoding="utf-8").read()
profiles = {}
for m in re.finditer(r'^  \["(?:[^"\\]|\\.)*", (\{.*\})\],$', profiles_src, re.M):
    p = json.loads(m.group(1))
    profiles[p["name"]] = p

def ts_entries(src, array_name):
    block = re.search(rf"export const {array_name}[^=]*= \[\n(.*?)\n\];", src, re.S)
    if not block:
        return []
    out = []
    for line in block.group(1).split("\n"):
        fields = {}
        for key in ("id", "animal", "type", "title", "date", "description", "provider", "notes", "drug", "dose"):
            fm = re.search(rf'{key}: ("(?:[^"\\]|\\.)*")', line)
            if fm:
                fields[key] = json.loads(fm.group(1))
        if fields:
            out.append(fields)
    return out

annual_exams = ts_entries(profiles_src, "annualExamEntries")
revised_medical = ts_entries(profiles_src, "revisedMedicalEntries")
trim_visits = ts_entries(profiles_src, "adoptionTrimVisits")
stats = {
    k: int(v)
    for k, v in re.findall(r"(\w+): (\d+),", re.search(r"sanctuaryStats[^{]*\{(.*?)\}", profiles_src, re.S).group(1))
}

dw_src = open("src/lib/deworming-vaccination-data.ts", encoding="utf-8").read()
dw_entries = ts_entries(dw_src, "importedDewormingEntries")
vx_entries = ts_entries(dw_src, "importedVaccinationEntries")
note_entries = ts_entries(dw_src, "checklistNoteEntries")
yard_rows = ts_entries(dw_src, "yardWideDewormings")
next_vacc = dict(
    re.findall(r'^  \["((?:[^"\\]|\\.)*)", "([\d-]+)"\],?$', re.search(r"nextVaccinationByAnimal[^=]*= new Map\(\[\n(.*?)\n\]\);", dw_src, re.S).group(1), re.M)
)

# ════════════════════ ADOPTION SHEET ════════════════════
wb = openpyxl.load_workbook(ADOPTION_XLSX, data_only=True)
ws = wb["Donkey Dreams - Current"]

sheet_count = 0
tot = dict(momBaby=0, bonded=0.0, special=0, over20=0, under3=0, chip=0)
totals_row = None

for row in ws.iter_rows(min_row=2):
    v = [c.value for c in row]
    if v[0] is None:
        if any(x is not None for x in v[1:7]):
            totals_row = v[1:7]
        continue
    name = resolve(str(v[0]))
    sheet_count += 1
    p = profiles.get(name)
    if not p:
        issue(f"[adoption] {name}: MISSING from donkeyProfiles")
        continue

    # identity fields
    checks = [
        ("herd", HERD_MAP.get(txt(v[7]), txt(v[7])), p["herd"]),
        ("sex", txt(v[8]), p["sex"]),
        ("size", txt(v[9]), p["size"]),
        ("color", txt(v[10]), p["color"]),
        ("intakeDate", iso(v[11]), p["intakeDate"]),
        ("birthDate", iso(v[13]), p["birthDate"]),
        ("origin", txt(v[14]), p["origin"]),
        ("lastAnnualExam", iso(v[19]), p["lastAnnualExam"]),
    ]
    for field, want, got in checks:
        if want != got and not (want is None and got is None):
            issue(f"[adoption] {name}.{field}: sheet={want!r} app={got!r}")

    # microchip: digits-only comparison
    chip_digits = re.sub(r"\D", "", txt(v[12]))
    app_digits = re.sub(r"\D", "", p["microchip"] or "")
    if chip_digits != app_digits:
        issue(f"[adoption] {name}.microchip: sheet digits={chip_digits!r} app={app_digits!r}")

    # flags
    mb = int(v[1]) if v[1] is not None else 0
    bonded = float(v[2]) if v[2] is not None else 0.0
    flags = [
        ("momBabyCount", mb, p["momBabyCount"]),
        ("isBondedPair", bonded > 0, p["isBondedPair"]),
        ("isSpecialNeeds", v[3] is not None, p["isSpecialNeeds"]),
        ("isOver20", v[4] is not None, p["isOver20"]),
        ("isUnder3", v[5] is not None, p["isUnder3"]),
        ("needsChip", v[6] is not None, p["needsChip"]),
    ]
    for field, want, got in flags:
        if want != got:
            issue(f"[adoption] {name}.{field}: sheet={want!r} app={got!r}")
    tot["momBaby"] += mb
    tot["bonded"] += bonded
    tot["special"] += 1 if v[3] is not None else 0
    tot["over20"] += 1 if v[4] is not None else 0
    tot["under3"] += 1 if v[5] is not None else 0
    tot["chip"] += 1 if v[6] is not None else 0

    # medical / special-needs text → entries
    med, sn = txt(v[17]), txt(v[18])
    descs = " || ".join(e["description"] for e in revised_medical if e["animal"] == name)
    if med and med.rstrip(".") not in descs:
        issue(f"[adoption] {name}: Medical text not found in entries: {med[:60]!r}")
    if sn and sn.rstrip(".") not in descs:
        issue(f"[adoption] {name}: Special Needs text not found in entries: {sn[:60]!r}")

    # annual exam entry
    exam = iso(v[19])
    if exam and not any(e["animal"] == name and e["date"] == exam for e in annual_exams):
        issue(f"[adoption] {name}: annual exam {exam} missing from annualExamEntries")

    # trim history
    trim = txt(v[20])
    if trim and trim.upper() != "N/A":
        if not any(t["animal"] == name for t in trim_visits):
            issue(f"[adoption] {name}: Trim History {trim[:50]!r} produced no trim visit")

if sheet_count != len(profiles):
    issue(f"[adoption] sheet has {sheet_count} donkeys, app has {len(profiles)} profiles")

expected_stats = dict(
    totalDonkeys=sheet_count, momBaby=tot["momBaby"], bondedPairs=round(tot["bonded"]),
    specialNeeds=tot["special"], seniors=tot["over20"], under3=tot["under3"], needsChip=tot["chip"],
)
for k, want in expected_stats.items():
    if stats.get(k) != want:
        issue(f"[adoption] sanctuaryStats.{k}: sheet-derived={want} app={stats.get(k)}")
if totals_row:
    tr = [int(float(x)) for x in totals_row]
    derived = [tot["momBaby"], int(tot["bonded"]), tot["special"], tot["over20"], tot["under3"], tot["chip"]]
    if tr != derived:
        issue(f"[adoption] sheet totals row {tr} != column sums {derived}")

ok.append(f"adoption: {sheet_count} donkeys, {len(annual_exams)} exams, {len(revised_medical)} medical texts, {len(trim_visits)} trims, stats {stats}")

# ════════════════════ CHECKLIST SHEET ════════════════════
wb2 = openpyxl.load_workbook(CHECKLIST_XLSX, data_only=True)
ws2 = wb2["Donkey Dreams - Current"]
HERDS = {"Angels", "Brave", "Dragons", "Elsie", "Legacy", "Pegasus", "Pink", "Senior", "Seniors", "Unicorns"}

def cell_dates(text, fallbacks):
    """All ISO dates mentioned in a history cell (year borrowed when absent)."""
    out = []
    tokens = list(re.finditer(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?", text))
    for i, m in enumerate(tokens):
        y = m.group(3)
        if not y:
            nxt = next((t.group(3) for t in tokens[i + 1:] if t.group(3)), None)
            y = nxt or (fallbacks[0][:4] if fallbacks and fallbacks[0] else None)
            if not y:
                continue
        if len(y) == 2:
            y = "20" + y
        out.append(f"{y}-{int(m.group(1)):02d}-{int(m.group(2)):02d}")
    return out

chk_count = 0
notes_count = 0
yard_sheet = set()
for row in ws2.iter_rows(min_row=2):
    v = [c.value for c in row]
    if v[0] is None:
        continue
    raw = str(v[0]).strip()
    herd = txt(v[1]) if len(v) > 1 else ""
    if raw.upper().startswith("PREV-HERD"):
        d, drug, dose = iso(v[1]), txt(v[2]), txt(v[3])
        if d and drug:
            yard_sheet.add((d, drug.upper(), dose))
        continue
    if herd not in HERDS:
        continue
    name = resolve(raw)
    chk_count += 1
    if name not in profiles:
        issue(f"[checklist] {name}: not in donkeyProfiles (roster mismatch)")

    dw_cell, vx_cell = txt(v[3]), txt(v[5])
    dew_fallback, vacc_fallback = iso(v[2]), iso(v[6])

    my_dw = {e["date"] for e in dw_entries if e["animal"] == name}
    my_vx = {e["date"] for e in vx_entries if e["animal"] == name} | my_dw

    for d in cell_dates(dw_cell, [dew_fallback]):
        if d not in my_dw:
            issue(f"[checklist] {name}: DW date {d} in cell but no deworming entry")
    for d in cell_dates(vx_cell, [vacc_fallback]):
        if d not in my_vx:
            issue(f"[checklist] {name}: VX date {d} in cell but no entry")

    # reverse: every entry date appears in the cell
    for e in dw_entries:
        if e["animal"] == name and e["date"] not in cell_dates(dw_cell, [dew_fallback]) and e["date"] not in cell_dates(vx_cell, [vacc_fallback]):
            issue(f"[checklist] {name}: entry {e['title']} {e['date']} not in sheet cell")

    nv = iso(v[7])
    if (nv or None) != next_vacc.get(name):
        issue(f"[checklist] {name}.nextVaccination: sheet={nv!r} app={next_vacc.get(name)!r}")

    if txt(v[8]):
        notes_count += 1
        note_text = txt(v[8])
        core = re.sub(r"^\d{1,2}/\d{1,2}/\d{2,4}\s*", "", note_text)
        if not any(e["animal"] == name and (core in e["description"] or note_text in e["description"]) for e in note_entries):
            issue(f"[checklist] {name}: note not imported: {note_text[:50]!r}")

yard_app = {(r["date"], r["drug"].upper(), r["dose"]) for r in yard_rows}
if yard_sheet != yard_app:
    issue(f"[checklist] yard-wide mismatch: sheet={sorted(yard_sheet)} app={sorted(yard_app)}")

if notes_count != len(note_entries):
    issue(f"[checklist] {notes_count} notes in sheet vs {len(note_entries)} imported")
if chk_count != len(next_vacc):
    pass  # some donkeys may lack next-vacc; per-row check above covers it

ok.append(f"checklist: {chk_count} donkeys, {len(dw_entries)} dw + {len(vx_entries)} vx entries, {len(note_entries)} notes, {len(next_vacc)} next-vacc, {len(yard_rows)} yard rows")

# ════════════════════ REPORT ════════════════════
print("\n".join(ok))
print()
if issues:
    print(f"DISCREPANCIES: {len(issues)}")
    for i in issues:
        print("  " + i)
else:
    print("DISCREPANCIES: 0 — both sheets fully reconciled ✔")
